from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from ..models import Receipt


HEADER = [
    "id",
    "store_name",
    "date",
    "total_amount",
    "category",
    "image_path",
    "created_at",
]


def _ensure_export_path(export_path: str) -> Path:
    p = Path(export_path)
    p.parent.mkdir(parents=True, exist_ok=True)
    return p


def append_receipt_to_excel(receipt: Receipt, export_path: str) -> str:
    path = _ensure_export_path(export_path)

    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
        if ws.max_row < 1:
            ws.append(HEADER)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "receipts"
        ws.append(HEADER)

    created_at = receipt.created_at
    if isinstance(created_at, datetime):
        created_at_str = created_at.isoformat()
    else:
        created_at_str = str(created_at)

    ws.append(
        [
            receipt.id,
            receipt.store_name,
            receipt.date,
            receipt.total_amount,
            receipt.category,
            receipt.image_path,
            created_at_str,
        ]
    )

    wb.save(path)
    return str(path)
